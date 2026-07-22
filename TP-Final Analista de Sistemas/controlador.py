import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog
from typing import Dict, List, Optional, Any
import gestor_pdf
import vistas
import config
from modelos import Usuario, Tarea
from logger import get_logger
from dao import UsuarioDAO, TareaDAO
from seguridad import Seguridad

logger = get_logger(__name__)

class Controlador:
    """
    Controlador principal MVC.
    Encapsula la lógica de negocio (Mock data, RBAC) y las funciones de archivo (PDF)
    que antes residían en main.py.
    """
    def __init__(self, root: Any) -> None:
        # root es la instancia de Aplicacion (tk.Tk)
        self.root = root
        self.usuario_actual: Optional[Usuario] = None

    def verificar_credenciales(self, usuario: str, contrasena: str) -> bool:
        """Verifica credenciales en SQL Server e inicia sesión."""
        try:
            user = UsuarioDAO.obtener_por_username(usuario)
            if user and Seguridad.verificar_password(contrasena, user.password_hash):
                self.usuario_actual = user
                logger.info(f"Usuario {usuario} autenticado exitosamente como {user.rol}.")
                
                # Auditoría
                Seguridad.registrar_auditoria(user.id_usuario, "Login exitoso")
                
                # Al loguearse, mostrar la barra de menú nativa
                self.root.config(menu=self.root.barra_menus)
                self.root.title(f"{config.TITULO_APP} - Dashboard")
                # Redirigir a Dashboard
                self.root.mostrar_frame(vistas.PaginaTareas)
                return True
                
            logger.warning(f"Intento fallido para el usuario: {usuario}")
            return False
        except Exception:
            logger.exception("Error al verificar credenciales en BD.")
            return False

    def cerrar_sesion(self) -> None:
        """
        Cierra la sesión del usuario actual.

        Limpia el estado de autenticación, oculta la barra de menú nativa
        y redirige a la pantalla de Login.
        """
        if self.usuario_actual:
            Seguridad.registrar_auditoria(
                self.usuario_actual.id_usuario, "Cerró sesión"
            )
            logger.info(f"Usuario {self.usuario_actual.username} cerró sesión.")
        self.usuario_actual = None
        self.root.config(menu="")
        self.root.title(config.TITULO_APP)
        self.root.mostrar_frame(vistas.PaginaLogin)

    def obtener_tareas_dashboard(self) -> List[Tarea]:
        """Obtiene tareas filtradas según el rol del usuario."""
        if not self.usuario_actual:
            return []
        if self.usuario_actual.rol == "JEFE/GERENTE":
            return TareaDAO.obtener_todas()
        else:
            return TareaDAO.obtener_tareas_por_usuario(self.usuario_actual.id_usuario)
            
    def obtener_metodistas(self) -> List[Usuario]:
        """Obtiene los usuarios con rol METODISTA."""
        return UsuarioDAO.obtener_metodistas()
        
    def obtener_opciones_estado(self) -> List[str]:
        """Devuelve las opciones de estado permitidas según el rol del usuario."""
        if not self.usuario_actual:
            return []
        if self.usuario_actual.rol == "JEFE/GERENTE":
            return ["Pendiente", "En Proceso", "Finalizado", "Cerrado"]
        else:
            return ["Pendiente", "En Proceso", "Finalizado"]
        
    def crear_tarea(self, titulo: str, descripcion: str, id_usuario_asignado: int) -> bool:
        """Delega la creación de la tarea al DAO."""
        exito = TareaDAO.crear_tarea(titulo, descripcion, id_usuario_asignado)
        if exito and self.usuario_actual:
            Seguridad.registrar_auditoria(self.usuario_actual.id_usuario, f"Creó tarea: {titulo}")
        return exito
        
    def borrar_tarea(self, id_tarea: int) -> bool:
        """Delega el borrado de la tarea al DAO."""
        exito = TareaDAO.borrar_tarea(id_tarea)
        if exito and self.usuario_actual:
            Seguridad.registrar_auditoria(self.usuario_actual.id_usuario, f"Borrador tarea ID: {id_tarea}")
        return exito
        
    def cambiar_estado_tarea(self, id_tarea: int, nuevo_estado: str) -> bool:
        try:
            exito = TareaDAO.actualizar_estado(id_tarea, nuevo_estado)
            if exito and self.usuario_actual:
                Seguridad.registrar_auditoria(self.usuario_actual.id_usuario, f"Cambio de estado en tarea {id_tarea} a {nuevo_estado}")
                logger.info(f"Tarea {id_tarea} cambió a estado {nuevo_estado}.")
            return exito
        except Exception:
            logger.exception("Error al cambiar estado de la tarea.")
            return False

    def registrar_nuevo_usuario(self, nombre: str, username: str, rol: str, contrasena: str) -> None:
        """Registra un usuario en el sistema previniendo escalada de privilegios."""
        if rol == "JEFE/GERENTE":
            master_pwd = simpledialog.askstring("Autorización Requerida", 
                                                "Ingrese la Contraseña Maestra para crear un Administrador:",
                                                show="*", parent=self.root)
            if master_pwd != config.MASTER_PASSWORD:
                messagebox.showerror("Acceso Denegado", "Contraseña maestra incorrecta. Operación cancelada.")
                return

        # Verificar si ya existe para dar feedback claro
        if UsuarioDAO.obtener_por_username(username) is not None:
            messagebox.showerror("Error", "El nombre de usuario ya está en uso.")
            return

        hash_pwd = Seguridad.generar_hash(contrasena)
        
        if UsuarioDAO.crear_usuario(username, hash_pwd, rol, nombre):
            messagebox.showinfo("Éxito", f"Usuario '{username}' ({rol}) creado correctamente.")
            if self.usuario_actual:
                Seguridad.registrar_auditoria(self.usuario_actual.id_usuario, f"Creó usuario: {username} ({rol})")
            self.root.mostrar_frame(vistas.PaginaLogin)
        else:
            messagebox.showerror("Error", "No se pudo crear el usuario en la base de datos.")

    # --- LÓGICA MOVIDA DESDE MAIN.PY ---

    def nueva_plantilla(self) -> None:
        """Renderiza la vista original preexistente para crear una nueva plantilla de cronometraje."""
        self.root.mostrar_frame(vistas.Plantilla)

    def abrir_archivo(self) -> None:
        """Abre un archivo utilizando filedialog."""
        tipos = [('Archivos de texto', '*.txt'), ('Todos los archivos', '*.*')]
        ruta = filedialog.askopenfilename(title="Seleccionar un archivo", filetypes=tipos)
        if ruta:
            messagebox.showinfo("Abrir", f"Archivo seleccionado: {ruta}")

    def guardar_archivo(self) -> bool:
        """
        Guarda el PDF con auto-naming basado en el código del artículo.

        Lee el código del artículo desde la vista Plantilla para generar
        un nombre de archivo predeterminado. Extrae los datos completos
        de la grilla y delega al gestor_pdf para el renderizado landscape.

        Returns:
            bool: True si el PDF se generó exitosamente, False en caso contrario.
        """
        # Fase 1: Auto-Naming — leer código del artículo desde la vista
        vista_plantilla = self.root.frames.get(vistas.Plantilla)
        nombre_articulo: str = "REPORTE_NUEVO"
        if vista_plantilla and hasattr(vista_plantilla, 'entry_articulo_codigo'):
            codigo = vista_plantilla.entry_articulo_codigo.get().strip()
            if codigo:
                nombre_articulo = codigo

        ruta = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Archivos PDF", "*.pdf")],
            initialfile=f"{nombre_articulo}.pdf"
        )
        if not ruta:
            return False

        try:
            if vista_plantilla and hasattr(vista_plantilla, 'obtener_datos_limpios'):
                datos_completos: Dict[str, Any] = vista_plantilla.obtener_datos_limpios()
                exito = gestor_pdf.generar_reporte_pdf(datos_completos, ruta)
                if exito:
                    messagebox.showinfo("Éxito", "PDF generado correctamente.")
                    return True
                else:
                    messagebox.showerror("Error", "Fallo al generar PDF. Revise el log del sistema.")
                    return False
            else:
                logger.error("No se encontró la vista Plantilla para obtener datos.")
                messagebox.showerror("Error", "La vista actual no soporta el guardado de datos.")
                return False
        except Exception:
            logger.exception("Error inesperado al intentar guardar el PDF.")
            messagebox.showerror("Error", "Ocurrió un error crítico. Revise el archivo de log.")
            return False

    def obtener_siguiente_id_relevamiento(self) -> int:
        """Obtiene el correlativo máximo + 1 de Relevamientos de la BD."""
        from dao import RelevamientoDAO
        return RelevamientoDAO.obtener_siguiente_id()

    def obtener_lista_maquinas(self) -> List[str]:
        """Obtiene la lista de nombres de máquinas registradas."""
        from dao import RecursoDAO
        return [r.nombre for r in RecursoDAO.obtener_por_tipo("MAQUINA")]

    def obtener_lista_hms(self) -> List[str]:
        """Obtiene la lista de nombres de herramientas/matrices registradas."""
        from dao import RecursoDAO
        return [r.nombre for r in RecursoDAO.obtener_por_tipo("HM")]

    def obtener_lista_operaciones(self) -> List[str]:
        """Obtiene la lista de nombres de operaciones registradas.

        Reutiliza la tabla Recursos almacenando las operaciones bajo
        el tipo 'OPERACION'. Retorna una lista de strings con los
        nombres disponibles para poblar el Combobox autocompletable
        del campo Operación en la Plantilla.

        Returns:
            Lista de nombres de operaciones registradas. Lista vacía
            si no existen registros.
        """
        from dao import RecursoDAO
        return [r.nombre for r in RecursoDAO.obtener_por_tipo("OPERACION")]

    def obtener_lista_operarios(self) -> List[str]:
        """Obtiene la lista de nombres de operarios registrados."""
        from dao import OperarioDAO
        return [o.nombre for o in OperarioDAO.obtener_todos()]

    def crear_recurso_silencioso(self, nombre: str, tipo: str) -> bool:
        """Crea silenciosamente un recurso (maquina o HM) si no existe."""
        from dao import RecursoDAO
        recurso = RecursoDAO.crear_silencioso(nombre, tipo)
        return recurso is not None

    def crear_recurso(self, nombre: str, tipo: str) -> bool:
        """
        Crea un nuevo recurso (MAQUINA o HM) de forma explícita en la base de datos.
        Asegura que el recurso no exista previamente antes de crearlo.

        Args:
            nombre (str): Nombre o descripción del recurso.
            tipo (str): Tipo del recurso ('MAQUINA' o 'HM').

        Returns:
            bool: True si el recurso se crea exitosamente o ya existía, False en caso de error.
        """
        from dao import RecursoDAO
        try:
            nombre_limpio = nombre.strip()
            if not nombre_limpio:
                logger.warning("Intento de crear un recurso con nombre vacío.")
                return False
            
            # Verificar existencia previa
            existente = RecursoDAO.obtener_por_descripcion_y_tipo(nombre_limpio, tipo)
            if existente:
                logger.info(f"El recurso '{nombre_limpio}' de tipo '{tipo}' ya existe.")
                return True
            
            recurso = RecursoDAO.crear_silencioso(nombre_limpio, tipo)
            if recurso:
                logger.info(f"Recurso '{nombre_limpio}' de tipo '{tipo}' creado explícitamente.")
                if self.usuario_actual:
                    from seguridad import Seguridad
                    Seguridad.registrar_auditoria(
                        self.usuario_actual.id_usuario,
                        f"Creó recurso explícito: {nombre_limpio} ({tipo})"
                    )
                return True
            return False
        except Exception:
            logger.exception(f"Error inesperado al crear recurso '{nombre}' del tipo '{tipo}'.")
            return False

    def borrar_recurso(self, nombre: str, tipo: str) -> bool:
        """
        Elimina un recurso (MAQUINA o HM) por nombre y tipo.

        Delega al DAO, registra auditoría si hay usuario autenticado,
        y captura excepciones de forma granular (Regla 7).

        Args:
            nombre: Descripción/nombre del recurso a eliminar.
            tipo: Tipo del recurso ('MAQUINA' o 'HM').

        Returns:
            True si la eliminación fue exitosa, False en caso contrario.
        """
        from dao import RecursoDAO
        try:
            nombre_limpio = nombre.strip()
            if not nombre_limpio:
                logger.warning("Intento de eliminar un recurso con nombre vacío.")
                return False
            exito = RecursoDAO.borrar_por_nombre_y_tipo(nombre_limpio, tipo)
            if exito and self.usuario_actual:
                Seguridad.registrar_auditoria(
                    self.usuario_actual.id_usuario,
                    f"Eliminó recurso: {nombre_limpio} ({tipo})"
                )
            return exito
        except Exception:
            logger.exception(f"Error inesperado al eliminar recurso '{nombre}' ({tipo}).")
            return False

    def borrar_relevamiento(self, id_relevamiento: int) -> bool:
        """
        Elimina un relevamiento completo (cabecera + tiempos) por ID.

        Delega al DAO, registra auditoría si hay usuario autenticado,
        y captura excepciones de forma granular (Regla 7).

        Args:
            id_relevamiento: ID del relevamiento a eliminar.

        Returns:
            True si la eliminación fue exitosa, False en caso contrario.
        """
        from dao import RelevamientoDAO
        try:
            exito = RelevamientoDAO.borrar_por_id(id_relevamiento)
            if exito and self.usuario_actual:
                Seguridad.registrar_auditoria(
                    self.usuario_actual.id_usuario,
                    f"Eliminó relevamiento ID: {id_relevamiento}"
                )
            return exito
        except Exception:
            logger.exception(f"Error inesperado al eliminar relevamiento {id_relevamiento}.")
            return False

    def crear_operario_silencioso(self, nombre: str) -> bool:
        """Crea silenciosamente un operario si no existe."""
        from dao import OperarioDAO
        operario = OperarioDAO.crear_silencioso(nombre)
        return operario is not None

    def guardar_relevamiento_bd(self, relevamiento: Any) -> bool:
        """
        Delega la persistencia del relevamiento transaccional al RelevamientoDAO.
        Aplica trazabilidad estricta y manejo de errores (Regla 7 y Prohibición de Silenciar Errores).
        """
        from dao import RelevamientoDAO
        from tkinter import messagebox
        import pprint
        
        # Trazabilidad requerida en consola
        print("DEBUG - Intentando guardar relevamiento. Estructura de datos:")
        try:
            pprint.pprint(vars(relevamiento))
        except Exception:
            print(relevamiento)
            
        try:
            exito = RelevamientoDAO.guardar(relevamiento)
            return exito
        except Exception as e:
            logger.exception("Error DB al intentar guardar el relevamiento.")
            messagebox.showerror("Error de Base de Datos", f"Se produjo un error crítico al guardar en la base de datos:\n\n{str(e)}")
            return False

    def abrir_relevamiento_desde_id(self, id_seleccionado: int) -> None:
        """Carga un relevamiento completo desde la BD y abre la vista Plantilla inyectando los datos."""
        print(f"DEBUG: Intentando cargar ID {id_seleccionado}")
        from dao import RelevamientoDAO
        import vistas
        relevamiento = RelevamientoDAO.obtener_por_id(id_seleccionado)
        if relevamiento:
            self.root.mostrar_frame(vistas.Plantilla)
            vista_plantilla = self.root.frames[vistas.Plantilla]
            id_tarea_asociada = relevamiento.id_tarea or 0
            vista_plantilla.cargar_datos(relevamiento, id_tarea_asociada)

    def abrir_relevamiento_desde_tarea(self, id_tarea: int) -> None:
        """
        Busca si la tarea tiene un relevamiento asociado.
        Si existe, lo carga y abre la vista Plantilla.
        Si no existe, abre una Plantilla limpia vinculada a la tarea.
        
        Args:
            id_tarea (int): El ID de la tarea del Treeview.
        """
        from dao import RelevamientoDAO
        import vistas
        try:
            relevamiento = RelevamientoDAO.obtener_por_tarea(id_tarea)
            # Cambiar de vista a Plantilla
            self.root.mostrar_frame(vistas.Plantilla)
            
            # Obtener la instancia de la vista Plantilla
            vista_plantilla = self.root.frames[vistas.Plantilla]
            
            if relevamiento:
                logger.info(f"Cargando relevamiento existente {relevamiento.id_relevamiento} para la tarea {id_tarea}.")
                vista_plantilla.cargar_datos(relevamiento, id_tarea)
            else:
                logger.info(f"Inicializando relevamiento nuevo para la tarea {id_tarea}.")
                vista_plantilla.cargar_datos_nuevos(id_tarea)
                
        except Exception:
            logger.exception(f"Error al intentar abrir el relevamiento de la tarea {id_tarea}.")
            from tkinter import messagebox
            messagebox.showerror("Error", f"No se pudo cargar el relevamiento asociado a la tarea {id_tarea}.")

    def obtener_resumen_relevamientos(self) -> List[tuple]:
        """Obtiene un resumen de todos los relevamientos desde el DAO."""
        from dao import RelevamientoDAO
        return RelevamientoDAO.obtener_todos_resumen()

    def buscar_relevamientos_historicos(self, filtros: Dict[str, str]) -> List[tuple]:
        """
        Busca relevamientos históricos aplicando filtros combinados.

        Recibe los filtros de la vista, delega la búsqueda al DAO
        y retorna los resultados. Los errores se registran vía logger (Regla 7).

        Args:
            filtros: Diccionario con claves opcionales ('articulo', 'operario',
                     'maquina', 'hm') mapeadas a cadenas de texto.

        Returns:
            Lista de tuplas con formato:
            (ID, Fecha, Artículo, Operación, Máquina, HM, Operario).
        """
        from dao import RelevamientoDAO
        try:
            return RelevamientoDAO.buscar_historico(filtros)
        except Exception:
            logger.exception("Error al buscar relevamientos históricos desde el controlador.")
            return []


    def importar_desde_excel(self, ruta_archivo: str) -> None:
        """
        Importa datos desde una planilla Excel (.xlsx/.xlsm) y los carga en la vista Plantilla.

        Realiza lectura completa de cabecera (operación en F2, descripción en F3,
        código artículo en B7, máquina/HM/operario/fecha en columna B) y escaneo
        dinámico de la grilla de tiempos a partir de la columna D (índice 4).

        El orden de inyección es crítico: primero se muestra el frame (disparando
        actualizar_vista/limpiar_datos) y luego se cargan los datos, evitando que
        el ciclo de vida de Tkinter borre la información recién inyectada.

        Args:
            ruta_archivo: Ruta absoluta al archivo Excel (.xlsx/.xlsm).

        Raises:
            No lanza excepciones al exterior; captura cualquier error y lo
            muestra al usuario vía messagebox, registrando el detalle en el log.
        """
        import sys
        import subprocess
        import datetime
        from tkinter import messagebox
        from modelos import Articulo, Recurso, Operario, Relevamiento, TiemposElemento

        try:
            # 1. Asegurar dependencias de forma silenciosa
            try:
                import openpyxl
            except ImportError:
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl

            # 2. Cargar hoja activa
            wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
            hoja = wb.worksheets[0]

            def safe_str(cell_ref: str) -> str:
                """Extrae el valor de una celda como string limpio, o cadena vacía si es None."""
                val = hoja[cell_ref].value
                return str(val).strip() if val is not None else ""

            def safe_float(val: Any, default: float = 0.0) -> float:
                """Convierte un valor de celda a float de forma segura, tolerando comas y '%'."""
                if val is None:
                    return default
                if isinstance(val, (int, float)):
                    return float(val)
                try:
                    return float(str(val).replace(',', '.').replace('%', '').strip())
                except ValueError:
                    return default

            # 3. Lectura de cabecera (coordenadas confirmadas por PoC)
            art_codigo = safe_str('A6')
            operacion = safe_str('F2')
            descripcion = safe_str('F3')
            maquina = safe_str('A20')
            hm = safe_str('A22')
            operario = safe_str('A26')

            # --- LECTURA DE FECHA Y NÚMERO DE CRONOMETRAJE ---
            fecha_val = hoja['A28'].value
            if isinstance(fecha_val, datetime.datetime):
                fecha_obj = fecha_val.date()
            elif isinstance(fecha_val, datetime.date):
                fecha_obj = fecha_val
            elif isinstance(fecha_val, str):
                try:
                    fecha_obj = datetime.datetime.strptime(fecha_val.strip(), "%d/%m/%Y").date()
                except ValueError:
                    fecha_obj = datetime.date.today()
            else:
                fecha_obj = datetime.date.today()

            # Guardamos el número de la celda A30 en una variable
            num_crono_excel = safe_str('A30')

            piezas_golpe = int(safe_float(hoja['E43'].value, 1.0))
            articulos_golpe = int(safe_float(hoja['E45'].value, 1.0))
            operarios_maq = int(safe_float(hoja['E47'].value, 1.0))

            # 4. Escaneo dinámico de la grilla de tiempos
            tiempos_lista: List[TiemposElemento] = []
            col = 4  # Inicio de matriz en Columna D
            columnas_vacias = 0

            while True:
                # Buscar nombre del elemento (filas 4-9, ignorando encabezados de sub-columna)
                elem_nombre: Optional[str] = None
                for fila_busqueda in range(4, 10):
                    val = hoja.cell(row=fila_busqueda, column=col).value
                    if val and str(val).strip() and str(val).strip().lower() not in ['v', 'min', 'm/seg', 'l']:
                        elem_nombre = str(val).strip()
                        break

                # Verificar si hay datos numéricos en la zona de observaciones (filas 11-30)
                tiene_datos = False
                for r in range(11, 31):
                    if (hoja.cell(row=r, column=col).value is not None or
                            hoja.cell(row=r, column=col + 1).value is not None or
                            hoja.cell(row=r, column=col + 2).value is not None):
                        tiene_datos = True
                        break

                # Tolerancia de 2 columnas vacías consecutivas antes de cortar
                if not elem_nombre and not tiene_datos:
                    columnas_vacias += 1
                    if columnas_vacias >= 2:
                        break
                    col += 3
                    continue

                columnas_vacias = 0

                # Auto-naming para elementos sin título en celdas combinadas
                if not elem_nombre:
                    idx_elemento = ((col - 4) // 3) + 1
                    elem_nombre = f"Elemento {idx_elemento}"

               # Elementos Medidos (fila 32), Lote Frec. (fila 33) y Suplemento (fila 38)
                el_medidos = int(safe_float(hoja.cell(row=32, column=col + 1).value, 1.0))
                lote_f = safe_float(hoja.cell(row=33, column=col + 1).value, 1.0)
                
                sup_raw = safe_float(hoja.cell(row=38, column=col + 1).value, 0.0)
                # FIX: Convertimos todo a decimal puro (0.01, 0.05) para evitar el doble escalado en vistas.py
                suplemento = sup_raw / 100.0 if sup_raw >= 1.0 else sup_raw

                # Recorrer observaciones (filas 11 a 30)
                for row_idx in range(11, 31):
                    v_val = hoja.cell(row=row_idx, column=col).value
                    min_val = hoja.cell(row=row_idx, column=col + 1).value
                    seg_val = hoja.cell(row=row_idx, column=col + 2).value

                    # Solo procesar si hay al menos un dato en la fila
                    if v_val is not None or min_val is not None or seg_val is not None:
                        # Valoración (normalizar factor decimal a porcentaje)
                        v_f = safe_float(v_val, 100.0)
                        if 0 < v_f <= 1.5:
                            v_f *= 100.0

                        # Minutos y segundos
                        m_f = safe_float(min_val, 0.0)
                        s_f = safe_float(seg_val, 0.0)

                        # Sincronización MIN ↔ SEG
                        if s_f == 0.0 and m_f > 0.0:
                            s_f = m_f * 60.0
                        elif m_f == 0.0 and s_f > 0.0:
                            m_f = s_f / 60.0

                        tiempos_lista.append(TiemposElemento(
                            elemento=elem_nombre,
                            valoracion=v_f,
                            minutos=m_f,
                            segundos=s_f,
                            elementos_medidos=el_medidos,
                            lote_frecuencial=lote_f,
                            suplemento=suplemento
                        ))

                # Avanzar al siguiente bloque de 3 columnas
                col += 3
                if col > 100:
                    break

            # 5. Construir el objeto Relevamiento (borrador nuevo)
            art_obj = Articulo(id_articulo="0", codigo=art_codigo, descripcion=descripcion) if art_codigo else None
            maq_obj = Recurso(id_recurso="0", nombre=maquina, tipo="MAQUINA") if maquina else None
            hm_obj = Recurso(id_recurso="0", nombre=hm, tipo="HM") if hm else None
            op_obj = Operario(id_operario="0", nombre=operario, legajo="") if operario else None

            relevamiento = Relevamiento(
                id_relevamiento=None,
                id_tarea=None,
                articulo=art_obj,
                operacion=operacion,
                recurso=maq_obj,
                hm=hm_obj,
                operario=op_obj,
                fecha=fecha_obj,
                postura="pie",
                piezas_por_golpe=piezas_golpe,
                articulos_por_golpe=articulos_golpe,
                operarios_maquina=operarios_maq,
                tiempos=tiempos_lista,
            )
            relevamiento.numero_importado = num_crono_excel
            

            # 6. Inyectar en la vista Plantilla
            # FIX CICLO DE VIDA TKINTER: primero mostrar (dispara limpiar_datos),
            # luego cargar datos para que no sean borrados por actualizar_vista().
            import vistas
            self.root.mostrar_frame(vistas.Plantilla)
            vista_plantilla = self.root.frames[vistas.Plantilla]
            vista_plantilla.cargar_datos(relevamiento, 0)

            messagebox.showinfo(
                "Importación Exitosa",
                "Planilla cargada correctamente como borrador nuevo."
            )

        except Exception as e:
            import traceback
            traceback.print_exc()
            logger.exception("Error crítico durante la importación de Excel.")
            messagebox.showerror(
                "Error de Importación",
                f"Hubo un error al leer el archivo Excel:\n\n{str(e)}"
            )

    
    def calcular_metricas_globales(self, suma_TE_columnas: float, piezas_por_golpe: float, articulos_por_golpe: float) -> Dict[str, Any]:
        """
        Realiza el cálculo en cascada de las métricas globales y de Epicor.
        Retorna un diccionario con los valores formateados a 3 decimales o vacíos en caso de error/división por cero.
        """
        res = {
            "te_operacion": 0.0,
            "cadencia_ph": 0.0,
            "cadencia_h1000": 0.0,
            "estandar_epicor": 0.0,
            "dotacion_epicor": 0.0
        }
        
        try:
            if piezas_por_golpe <= 0.0 or articulos_por_golpe <= 0.0 or suma_TE_columnas <= 0.0:
                return res

            # TE_Operacion = suma_TE_columnas / piezas_por_golpe
            te_operacion = suma_TE_columnas / piezas_por_golpe
            res["te_operacion"] = te_operacion

            # cadencia_pcs_hr = 60 / TE_Operacion
            if te_operacion <= 0.0:
                return res
            cadencia_ph = 60.0 / te_operacion
            res["cadencia_ph"] = cadencia_ph

            # cadencia_hrs_1000 = 1000 / cadencia_pcs_hr
            if cadencia_ph <= 0.0:
                return res
            cadencia_h1000 = 1000.0 / cadencia_ph
            res["cadencia_h1000"] = cadencia_h1000

            # estandar_epicor = cadencia_hrs_1000
            estandar_epicor = cadencia_h1000
            res["estandar_epicor"] = estandar_epicor

            # dotacion_epicor = estandar_epicor / articulos_por_golpe
            dotacion_epicor = estandar_epicor / articulos_por_golpe
            res["dotacion_epicor"] = dotacion_epicor

        except ZeroDivisionError:
            logger.warning("ZeroDivisionError detectado y manejado silenciosamente en calcular_metricas_globales.")
        except Exception:
            logger.exception("Error inesperado en calcular_metricas_globales.")
            
        return res

    def crear_backup_bd(self) -> None:
        """
        Solicita al usuario un directorio de destino y ejecuta un backup
        nativo de SQL Server (BACKUP DATABASE) a través de ``database.ejecutar_backup``.

        El nombre del archivo se genera automáticamente con un timestamp
        (formato ``YYYYMMDD_HHMMSS``) para evitar colisiones. Las barras
        diagonales se normalizan a backslash porque SQL Server solo acepta
        rutas Windows nativas.

        Muestra un ``messagebox.showinfo`` con la ruta del backup si tiene
        éxito, o un ``messagebox.showerror`` si la operación falla.

        Returns:
            None. La función comunica el resultado directamente al usuario
            a través de diálogos modales de tkinter.
        """
        import datetime
        import database

        # 1. Solicitar directorio de destino al usuario
        directorio: str = filedialog.askdirectory(
            title="Seleccione la carpeta para guardar el Backup"
        )
        if not directorio:
            logger.info("Operación de backup cancelada por el usuario.")
            return

        # 2. Construir nombre de archivo con timestamp único
        fecha_str: str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_completa: str = f"{directorio}/Backup_MetodosTiempos_{fecha_str}.bak"

        # 3. Normalizar barras: SQL Server requiere rutas Windows con backslash
        ruta_completa = ruta_completa.replace("/", "\\")

        logger.info("Iniciando backup de base de datos hacia: %s", ruta_completa)

        # 4. Ejecutar el backup delegando a la capa de datos
        exito: bool = database.ejecutar_backup(ruta_completa)

        # 5. Feedback al usuario
        if exito:
            messagebox.showinfo(
                "Backup Exitoso",
                f"El backup de la base de datos se completó correctamente.\n\n"
                f"Ubicación:\n{ruta_completa}"
            )
            if self.usuario_actual:
                Seguridad.registrar_auditoria(
                    self.usuario_actual.id_usuario,
                    f"Backup de BD creado en: {ruta_completa}"
                )
        else:
            messagebox.showerror(
                "Error de Backup",
                "No se pudo completar el backup de la base de datos.\n\n"
                "Revise el archivo de log para más detalles."
            )

    def blanquear_credenciales(self, nombre_completo: str, clave_maestra: str, nueva_pwd: str) -> bool:
        """
        Blanquea (resetea) la contraseña de un usuario identificado por su nombre completo.

        Flujo de validación:
        1. Verifica la clave maestra contra ``config.MASTER_PASSWORD``.
        2. Busca al usuario en la BD por nombre completo exacto via DAO.
        3. Genera un nuevo hash bcrypt y actualiza la BD.
        4. Informa al usuario su username recuperado y registra la acción en auditoría.

        Este método está diseñado para ejecutarse desde la pantalla de Login,
        donde no existe sesión activa. La auditoría se registra con el ID del
        usuario modificado (no del usuario logueado, que no existe en este contexto).

        Args:
            nombre_completo: Nombre completo del usuario a recuperar (coincidencia exacta).
            clave_maestra: Contraseña maestra ingresada para autorizar la operación.
            nueva_pwd: Nueva contraseña en texto plano que será hasheada con bcrypt.

        Returns:
            True si la contraseña fue actualizada exitosamente, False en caso contrario.
        """
        # --- Validación de clave maestra ---
        if clave_maestra != config.MASTER_PASSWORD:
            messagebox.showerror(
                "Acceso Denegado",
                "Contraseña maestra incorrecta. Operación cancelada."
            )
            logger.warning("Intento de blanqueo de credenciales con clave maestra incorrecta.")
            return False

        # --- Búsqueda del usuario por nombre completo ---
        usuario: Optional[Usuario] = UsuarioDAO.obtener_por_nombre_completo(nombre_completo)
        if usuario is None:
            messagebox.showerror(
                "Usuario No Encontrado",
                f"No se encontró ningún usuario con el nombre exacto:\n'{nombre_completo}'.\n\n"
                f"Verifique que el nombre esté escrito correctamente (respetando mayúsculas y acentos)."
            )
            logger.warning(f"Blanqueo fallido: no se encontró usuario con nombre '{nombre_completo}'.")
            return False

        # --- Generación de hash y actualización ---
        try:
            nuevo_hash: str = Seguridad.generar_hash(nueva_pwd)
            exito: bool = UsuarioDAO.actualizar_password(usuario.id_usuario, nuevo_hash)

            if exito:
                messagebox.showinfo(
                    "Recuperación Exitosa",
                    f"Operación exitosa.\n\n"
                    f"Su usuario recuperado es: {usuario.username}\n"
                    f"La contraseña ha sido actualizada."
                )
                # Auditoría con el ID del usuario modificado (no hay sesión activa)
                Seguridad.registrar_auditoria(
                    usuario.id_usuario,
                    f"Blanqueo de contraseña ejecutado para '{usuario.username}'"
                )
                logger.info(f"Blanqueo de credenciales exitoso para usuario '{usuario.username}' (ID={usuario.id_usuario}).")
                return True
            else:
                messagebox.showerror(
                    "Error de Actualización",
                    "No se pudo actualizar la contraseña en la base de datos.\n"
                    "Revise el archivo de log para más detalles."
                )
                return False
        except Exception:
            logger.exception(f"Error crítico durante el blanqueo de credenciales para '{nombre_completo}'.")
            messagebox.showerror(
                "Error Crítico",
                "Ocurrió un error inesperado durante la operación.\n"
                "Revise el archivo de log para más detalles."
            )
            return False

